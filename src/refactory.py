import pandas as pd
import os
import shutil
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import create_engine
from colorama import Fore
from database import ConnectPostgresQL, OrdersTable
from datetime import datetime
from time import sleep
import locale
import glob 

class ExtractPipeline:
    
    def __init__(self, host):
        self.db_connection = ConnectPostgresQL(host)
        self.session = self.db_connection.Session()


    def list_files_in_directory(self, extractor_file_path, file_path_error):
        files_data = []
        try:
            for filename in os.listdir(extractor_file_path):
                if filename.endswith('.xlsx') and not filename.startswith('~$'):
                    xls_file = pd.ExcelFile(os.path.join(extractor_file_path, filename))
                    if '2-Resultado' in xls_file.sheet_names:
                        data = pd.read_excel(os.path.join(extractor_file_path, filename),
                                            sheet_name='2-Resultado', engine='openpyxl', header=1)
                        files_data.append((filename, data))
                    else:
                        # print(f"A folha '2-Resultado' não encontrada no arquivo: {filename}")
                        shutil.move(os.path.join(extractor_file_path, filename), os.path.join(file_path_error, filename))
                        # print(f'Arquivo {filename} movido para a pasta de arquivos com erros...')
                        
            print(f'Arquivos encontrados: {len(files_data)}')
        except FileNotFoundError:
            print('Diretório não encontrado...')
        except PermissionError:
            print('Sem permissão para acessar o diretório...')
        except Exception as e:
            print(f'Erro inesperado: {e}')
        
        return files_data
        
    
    def list_files_in_raw_directory(self, extractor_file_path):
        files_data = []
        try:
            for filename in os.listdir(extractor_file_path):
                if filename.endswith('.xlsx') and not filename.startswith('~$'):
                    file_path = os.path.join(extractor_file_path, filename)
                    extract_file = pd.read_excel(file_path, sheet_name='RELATÓRIO', engine='openpyxl')
                    files_data.append((filename, extract_file))
                else:
                    print('Arquivo não suportado...')
                    continue                
            return files_data
        
        except FileNotFoundError:
            print('Diretório não encontrado...')
        except PermissionError:
            print('Sem permissão para acessar o diretório...')
        except Exception as e:
            print(f'Erro inesperado: {e}')

    
    def verify_column(self, extractor_file_path, df_list, column_name, file_path_error):

        try:
            for filename, df in df_list:
                if column_name in df.columns:
                    return True

                else:
                    print(f'Coluna Pedido Faturamento não encontrada no arquivo {filename}...')
                    # move o arquivo para a pasta de erro
                    new_filename = f"{os.path.splitext(filename)[0]}_falta_coluna_pedido.xlsx"
                    # pega o caminho completo do arquivo com erro
                    old_file_path = os.path.join(extractor_file_path, filename)
                    # pega o caminho completo da pasta de erro
                    error_path = os.path.join(file_path_error, new_filename)
                    # cria a pasta de erro se não existir            
                    os.makedirs(file_path_error, exist_ok=True)                    
                    # move o arquivo para a pasta de erro
                    shutil.move(old_file_path, error_path)
                    print(f'Arquivo {filename} movido para a pasta de arquivos com erros...')
                    
        except FileNotFoundError:
            print('Arquivo não encontrado...')
        except PermissionError:
            print('Sem permissão para acessar o arquivo...')
        except Exception as e:
            print(f'Erro inesperado em verificar colunas: {e}')
       
    
    def identify_new_orders(self, df_list, column_name):
        new_orders_list = []
        try:
            for filename, df in df_list:
                if column_name in df.columns:
                    df[column_name] = pd.to_numeric(df[column_name], errors='coerce')
                    print(f'Total de registros no extrator {filename}: {len(df)}')

                    # Filtra apenas pedidos recentes
                    recent_orders = df[df[column_name] > 0]
                    # Remove valores nulos
                    recent_orders = recent_orders[recent_orders[column_name].notna()]
                    # Define novamente o tipo da coluna como inteiro
                    recent_orders[column_name] = recent_orders[column_name].astype(int)

                    # Consulta o banco de dados com os pedidos já inseridos anteriormente
                    existing_orders = set(int(order) for order in pd.read_sql_query(
                        f'SELECT DISTINCT pedido_faturamento FROM {OrdersTable.__tablename__}',
                        self.db_connection.engine)['pedido_faturamento'])

                    # Compara os pedidos do extrator e identifica os novos
                    new_orders_file = set(recent_orders[column_name]) - existing_orders
                    print(f'Novos pedidos encontrados no arquivo {filename}: {len(new_orders_file)}')

                    # Cria um DataFrame apenas com os novos pedidos
                    new_orders_df = recent_orders[recent_orders[column_name].isin(new_orders_file)]

                    # Adiciona o novo DataFrame à lista de novos pedidos, se não estiver vazio
                    if not new_orders_df.empty:
                        new_orders_list.append(new_orders_df)
                else:
                    print(f'Coluna {column_name} não encontrada no arquivo {filename}...')

            return new_orders_list

        except FileNotFoundError:
            print(Fore.RED + f'Arquivo não encontrado' + Fore.RESET)
        except PermissionError:
            print(Fore.RED + f'Sem permissão para acessar o arquivo' + Fore.RESET)
        except Exception as e:
            print(Fore.RED + f'Erro inesperado: {e}' + Fore.RESET)
                
        
    def columns_keeper(self):
        columns = ['CODIGO CLIENTE', 'NOME DO CLIENTE', 'LOJA CLIENTE', 'CNPJ DO CLIENTE', 'CNPJ DE FATURAMENTO',
                    'PROJETO', 'OBRA', 'ID EQUIPAMENTO', 'EQUIPAMENTO', 'DESCRICAO DO PRODUTO', 'DATA DE ATIVACAO LEGADO', 
                    'PERIODO DE FATURAMENTO', 'DIAS DE LOCACAO', 'VALOR UNITARIO', 'VALOR BRUTO', 'DATA DE ATIVACAO', 'QUANTIDADE', 
                    'VLR. TOTAL PEDIDO', 'VLR. TOTAL FATURAMENTO',
                    'NF DE FATURAMENTO',  'DATA DE FATURAMENTO', 'DATA BASE REAJUSTE', 'VALOR DE ORIGEM', 
                    'INDEXADOR', 'CALCULO REAJUSTE', 'INDICE APLICADO', 'ACRESCIMO', 'CONTRATO LEGADO', 'PEDIDO FATURAMENTO', 
                    'SERIE DO EQUIPAMENTO']
        
        
        return columns
    
    
    def rename_columns(self):
        columns = {'VLR. TOTAL PEDIDO': 'VALOR TOTAL GERADO', 'VLR. TOTAL FATURAMENTO': 'VALOR TOTAL FATURAMENTO'}
        return columns

    
    def create_files_with_new_orders(self, df_list, data_raw_path, column_name):
        try:
            
            if df_list:
                for df in df_list:
                    if not df.empty:  # Verifica se o DataFrame não está vazio
                        os.makedirs(data_raw_path, exist_ok=True)

                        if 'Nome do Cliente' in df.columns:
                            for order_number, order_group in df.groupby(column_name):
                                client_name_valid = order_group['Nome do Cliente'].iloc[0]\
                                    .translate(str.maketrans('.', ' ', r'\/:*?"<>|'))
                                filename = f'{order_number}_{client_name_valid}.xlsx'
                                file_path = os.path.join(data_raw_path, filename)
                                # salva o arquivo com cabeçalho em maiúsculas
                                order_group.columns = map(str.upper, order_group.columns)
                                # seleciona apenas as colunas necessárias
                                order_group = order_group[self.columns_keeper()]
                                # renomeia as colunas
                                order_group.rename(columns=self.rename_columns(), inplace=True)                                
                                
                                # formatação das colunas
                                TransformPipeline().format_columns_date([(filename, order_group)])
                                TransformPipeline().format_columns_cnpj([(filename, order_group)])

                                # salva o arquivo
                                order_group.to_excel(file_path, sheet_name='RELATÓRIO', index=False)

                        else:
                            print('Coluna Nome do Cliente não encontrada...')
            else:
                print('Nenhum novo pedido encontrado...')
                return False
        except FileNotFoundError:
            print('Arquivo não encontrado...')
        except PermissionError:
            print('Sem permissão para acessar o arquivo...')
        except Exception as e:
            print(f'Erro inesperado: {e}')

    
    def standard_columns_name(self, df_list):
        try:
            
            for df in df_list:
                if not df.empty:
                    df.columns = df.columns.str.lower().str.replace(' ', '_')\
                        .str.replace('.', '').str.replace('-', '').str.replace('ç', 'c').str.replace('?', '')                    
            print(f'Arquivos padronizados com sucesso...')

        except FileNotFoundError:
            print('Diretório não encontrado...')
        except PermissionError:
            print('Sem permissão para acessar o diretório...')
        except Exception as e:
            print(f'Erro inesperado em padronizar colunas: {e}')

                    
    def update_database(self, df_list):
        try:
            # Inicia a conexão com o banco de dados
            engine = self.db_connection.engine
            conn = engine.connect()
                    
            for df in df_list:
                if not df.empty:
                    df.to_sql(OrdersTable.__tablename__, conn, if_exists='append', index=False, chunksize=1000)
            
            print(f'Novos pedidos salvos no banco de dados.')
        except Exception as e:
            print(e)


    def add_new_columns_to_database(self, df_list):
        try:
            if df_list:
                all_columns = set()
                for df in df_list:
                    all_columns.update(df.columns)

                table_columns = set(OrdersTable.__table__.columns.keys())

                for column_name in all_columns:
                    if column_name not in table_columns:
                        try:
                            self.db_connection.engine.execute(
                                f'ALTER TABLE {OrdersTable.__tablename__} ADD COLUMN "{column_name}" TEXT'
                            )
                            print(Fore.CYAN + f'Coluna {column_name} adicionada com sucesso...' + Fore.RESET)
                        except Exception as e:
                            continue       
                    
            else:
                print('Nenhum arquivo encontrado...')
        except FileNotFoundError:
            print('Arquivo não encontrado...')
        except PermissionError:
            print('Sem permissão para acessar o arquivo...')
        except Exception as e:
            print(f'Erro inesperado: {e}')


class TransformPipeline:


    def format_columns_values(self, df_list, columns_to_format=['VALOR TOTAL GERADO', 'VALOR TOTAL FATURAMENTO']):
        """Função para tratar os dados das colunas especificadas,
        passando de strings para floats no formato que o Python
        entende como números para soma, evitando cálculo incorreto 
        do arquivo consolidado."""
        try:
            for filename, df in df_list:
                if isinstance(df, pd.DataFrame):
                    for col in columns_to_format:
                        if col in df.columns:
                            df[col] = df[col].astype(str).str.replace('.', '').str.replace(',', '.')
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                        else:
                            print(f'Coluna {col} não encontrada no arquivo {filename}...')
                    print(f'Colunas formatadas com sucesso no arquivo {filename}...')
                else:
                    print(f"Erro: O objeto associado ao arquivo {filename} não é um DataFrame.")
        except Exception as e:
            print(f'Erro ao formatar colunas: {e}')
                    

    def format_columns_cnpj(self, df_list, columns_to_format=['CNPJ DO CLIENTE', 'CNPJ DE FATURAMENTO']):
        try:
            for filename, df in df_list:
                if isinstance(df, pd.DataFrame):
                    for col in columns_to_format:
                        if col in df.columns:
                            df[col] = df[col].astype(str).str.replace(r'(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})', r'\1.\2.\3/\4-\5')
                            print(f'Colunas de CNPJ formatadas com sucesso no arquivo {filename}...')
                else:
                    print(f"Erro: O objeto associado ao arquivo {filename} não é um DataFrame.")
        except Exception as e:
            print(f'Erro ao formatar CNPJ: {e}')


    def format_columns_date(self, df_list, columns_to_format=['DATA DE ATIVACAO', 'DATA DE FATURAMENTO', 
                                                              'DATA BASE REAJUSTE', 'DATA DE ATIVACAO LEGADO']):
        try:
            for filename, df in df_list:
                if isinstance(df, pd.DataFrame):
                    for col in columns_to_format:
                        if col in df.columns:
                            # Verifica se a coluna já é do tipo datetime64, se não, converte
                            if not pd.api.types.is_datetime64_any_dtype(df[col]):
                                df[col] = pd.to_datetime(df[col], errors='coerce')
                            df[col] = df[col].dt.strftime('%d/%m/%Y')
                            print(f'Coluna de data {col} formatada com sucesso no arquivo {filename}...')
                        else:
                            print(f'Coluna {col} não encontrada no arquivo {filename}.')
                else:
                    print(f"Erro: O objeto associado ao arquivo {filename} não é um DataFrame.")
        except Exception as e:
            print(f'Erro ao formatar data: {e}')           
       
        
    def format_styles_report_sheet(self, df_list, directory):
        try:
            for filename, df in df_list:
                if isinstance(df, pd.DataFrame):
                    # Aplicar a lógica de conversão na coluna 'VALOR TOTAL FATURAMENTO' e 'VALOR TOTAL GERADO'
                    # df['VALOR TOTAL FATURAMENTO'] = df['VALOR TOTAL FATURAMENTO'].apply(self.corrigir_valor_faturamento)
                    # df['VALOR TOTAL GERADO'] = df['VALOR TOTAL GERADO'].apply(self.corrigir_valor_faturamento)

                    # Converte as colunas para números
                    numeric_columns = ['VALOR TOTAL FATURAMENTO', 'VALOR TOTAL GERADO']
                    for col in numeric_columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce')

                    # Caminho completo do arquivo
                    file_path = os.path.join(directory, filename)
                    
                    # Carrega o arquivo Excel existente
                    wb = load_workbook(file_path)
                    ws = wb['RELATÓRIO']

                    # Formatação das células
                    for column in range(1, ws.max_column + 1):
                        col_letter = ws.cell(row=1, column=column).column_letter
                        ws.column_dimensions[col_letter].width = 25
                        cell = ws.cell(row=1, column=column)
                        cell.font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        ws.row_dimensions[1].height = 24

                    # Salva as alterações no arquivo
                    wb.save(file_path)

                    print(f'Planilha {filename} formatada com sucesso.')
                else:
                    print(f"Erro: O objeto associado ao arquivo {filename} não é um DataFrame.")
        except Exception as e:
            print(f'Erro inesperado: {e}')


    def format_billing_values(self, df_list_with_names, columns_to_format=['VALOR TOTAL GERADO', 'VALOR TOTAL FATURAMENTO']):
        """Função para tratar os dados das colunas especificadas,
        passando de strings para floats no formato que o Python
        entende como números para soma, evitando cálculo incorreto 
        do arquivo consolidado."""
        for filename, df in df_list_with_names:
            if isinstance(df, pd.DataFrame):
                for col in columns_to_format:
                    if col in df.columns:
                        df[col] = df[col].astype(str).str.replace('.', '').str.replace(',', '.')
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                    else:
                        print(f'Coluna {col} não encontrada no arquivo {filename}...')
                print(f'Colunas formatadas com sucesso no arquivo {filename}...')
            else:
                print(f"Erro: O objeto associado ao arquivo {filename} não é um DataFrame.")
            try:
                if isinstance(valor, str):
                    valor = valor.replace('R$', '').replace('.', '').replace(',', '.')
                    return valor
                else:
                    return valor
            except Exception as e:
                print(f'Erro inesperado valor: {e}')
    
    
    def generate_synthesis_sheet(self, df_list, directory):
        try:
            for filename, df in df_list:
                if isinstance(df, pd.DataFrame):
                    # verifica se já existe a planilha 'Síntese'
                    if 'SÍNTESE' in pd.ExcelFile(os.path.join(directory, filename)).sheet_names:
                        print(f'Planilha "SÍNTESE" já existe em {filename}.')
                    else:
                        

                        # Criar a planilha 'Síntese'
                        sintese_df = df.groupby(['PROJETO', 'OBRA', 'CONTRATO LEGADO'], as_index=False).agg({'VALOR TOTAL GERADO': 'sum', 'VALOR TOTAL FATURAMENTO': 'sum'})
                        sintese_df = sintese_df.rename(columns={'VALOR TOTAL FATURAMENTO': 'VALOR TOTAL FATURADO'})
                        
                        # Caminho completo do arquivo
                        file_path = os.path.join(directory, filename)
                        
                        # Carrega o arquivo Excel existente
                        wb = load_workbook(file_path)
                        
                        # Adiciona a planilha 'Síntese'
                        wb.create_sheet('SÍNTESE')
                        ws = wb['SÍNTESE']
                        

                        # Escreve os cabeçalhos
                        headers = sintese_df.columns
                        for col_idx, header in enumerate(headers, start=1):
                            ws.cell(row=1, column=col_idx, value=header)
                            ws.cell(row=1, column=col_idx).font = Font(bold=True)
                        
                        # Escreve os dados da 'Síntese' na planilha
                        for r_idx, row in sintese_df.iterrows():
                            for c_idx, value in enumerate(row, start=1):
                                ws.cell(row=r_idx + 2, column=c_idx, value=value)
                        
                        # Salva as alterações no arquivo
                        wb.save(file_path)

                        print(f'Planilha "SÍNTESE" criada com sucesso em {filename}.')
                else:
                    print(f"Erro: O objeto associado ao arquivo {filename} não é um DataFrame.")
        except Exception as e:
            print(f'Erro inesperado: {e}')


    def format_styles_synthesis_sheet(self, df_list, data_raw_path):
        try:
            for filename, df in df_list:
                if isinstance(df, pd.DataFrame):
                    # Caminho completo do arquivo
                    file_path = os.path.join(data_raw_path, filename)
                    
                    # Carrega o arquivo Excel existente
                    wb = load_workbook(file_path)
                    ws = wb['SÍNTESE']

                    # Formatação das células
                    for column in range(1, ws.max_column + 1):
                        col_letter = ws.cell(row=1, column=column).column_letter
                        ws.column_dimensions[col_letter].width = 25
                        cell = ws.cell(row=1, column=column)
                        cell.font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        ws.row_dimensions[1].height = 24
                    
                    """ for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                                 top=Side(border_style='thin'), bottom=Side(border_style='thin')) """

                    # Salva as alterações no arquivo
                    wb.save(file_path)
                    print(f'Planilha "SÍNTESE" formatada com sucesso em {filename}.')
                else:
                    print(f"Erro: O objeto associado ao arquivo {filename} não é um DataFrame.")
        except Exception as e:
            print(f'Erro inesperado: {e}')  


class LoadPipeline:

    def move_files_to_month_subfolder(self, directory_origin, target_directory):

        if not any(file.endswith('.xlsx') for file in os.listdir(directory_origin)):
            print('Nenhum arquivo encontrado...')
            return False    
        
        # obtém os arquivos xlsx no subdiretório principal
        files_to_move = [file for file in os.listdir(directory_origin) if file.endswith('.xlsx')]
        
        # cria a subpasta do mês e ano
        current_date = datetime.now()
        # formata a data atual para o formato mm-aaaa
        month_year = current_date.strftime('%m-%Y')
        
        for file_to_move in files_to_move:

            # estabelece caminho completo do arquivo na origem
            current_file_path = os.path.join(directory_origin, file_to_move)
                            
            # extrai o nome do cliente do nome do arquivo
            client_name_start = file_to_move.find('_') + 1
            client_name_end = file_to_move.find('.', client_name_start)
            client_name = file_to_move[client_name_start:client_name_end]
            client_name = file_to_move[client_name_start:client_name_end].strip()
            print(f'Nome do cliente: {client_name}')               

            # estabelece caminho completo do arquivo na destino
            current_file_path_with_month = os.path.join(target_directory, client_name, month_year)
            print(f'Caminho do arquivo de destino: {current_file_path_with_month}')
            
            # cria o diretório para o arquivo ser movido
            if not os.path.exists(current_file_path_with_month):
                # se não existir, cria o diretório
                os.makedirs(current_file_path_with_month, exist_ok=True)
                print(f'Pasta {current_file_path_with_month} criada com sucesso!')    
                # move o arquivo para o diretório correspondente ao nome do cliente

            # caminho completo do arquivo de destino
            destination_file_path = os.path.join(current_file_path_with_month, file_to_move)
            
            # verifica se o arquivo já existe e remove o arquivo no caso positivo
            if os.path.exists(destination_file_path):
                print(f'Arquivo {file_to_move} já existe no diretório {current_file_path_with_month}')
                try:                
                    # os.system(f'taskkill /f /im EXCEL.EXE')
                    os.remove(current_file_path)
                except PermissionError as e:
                    continue

                print(f'Arquivo {file_to_move} removido do diretório de destino')
                
            # move o arquivo para o diretório correspondente ao nome do cliente no caso negativo
            try:
                # Tenta abrir o arquivo para evitar o erro WinError 32
                with open(current_file_path, 'rb') as f:
                    # Fecha o arquivo imediatamente
                    f.close()

                # Tenta mover o arquivo
                shutil.move(current_file_path, destination_file_path)
                print(f'Arquivo {file_to_move} movido para {current_file_path_with_month}')

            except PermissionError as e:
                with open(current_file_path, 'rb') as f:
                    f.close()
            
            except FileNotFoundError as e:
                print(f'O arquivo {file_to_move} não existe mais na origem: {e}')
                continue
            
            except OSError as e:

                print(f'Erro ao mover o arquivo {file_to_move} possívelmente aberto: {e}')
                return False
            

class ConsolidatePipeline:    
    
    def remove_file(self, folder_path):
        try:
            # remove todo arquivo que inicia com "CONSOLIDADO"
            for file in os.listdir(folder_path):
                if file.startswith('CONSOLIDADO'):
                    os.remove(os.path.join(folder_path, file))
                    sleep(0.5)
                    
        except FileNotFoundError:
            pass  

    
    def list_files_to_consolidate(self, input_path):
        self.remove_file(input_path)
        
        files_data = []
        # Obtenha o ano e mês atuais
        current_date = datetime.now()
        month_year = current_date.strftime('%m-%Y')

        if not os.path.exists(input_path):
            print('Diretório não encontrado...')
            return files_data
        
        # Itera sobre as pastas de clientes dentro do diretório principal
        for client_folder in os.listdir(input_path):
            client_folder_path = os.path.join(input_path, client_folder)
            # Verifica se é um diretório
            if os.path.isdir(client_folder_path):
                # Caminho para a pasta do mês atual dentro do cliente
                month_path = os.path.join(client_folder_path, month_year)
                # Verifica se o diretório do mês existe dentro do cliente
                if os.path.exists(month_path):
                    # Itera sobre os arquivos .xlsx dentro do diretório do mês
                    for file_name in os.listdir(month_path):
                        if file_name.endswith('.xlsx') and not file_name.startswith('~$'):
                            file_path = os.path.join(month_path, file_name)
                            consolidate_file = pd.read_excel(file_path, sheet_name='RELATÓRIO', engine='openpyxl')
                            files_data.append((file_name, consolidate_file))
        
        if len(files_data) == 1:
            print(f'Apenas 1 arquivo encontrado no diretório {month_path}...')        
        return files_data
    
    
    def merge_excel_reports(self, folder_path):
        # Obtenha o ano e mês atuais
        current_date = datetime.now()
        month_year = current_date.strftime('%m-%Y')

        try:
            # Iterar sobre os diretórios dentro do folder_path
            for subfolder in os.listdir(folder_path):
                subfolder_path = os.path.join(folder_path, subfolder)
                if os.path.isdir(subfolder_path):
                    month_folder_path = os.path.join(subfolder_path, month_year)
                    if os.path.exists(month_folder_path):
                        excel_files = [files for files in os.listdir(month_folder_path) if files.endswith('.xlsx') and not files.startswith('~$')]
                        
                        if len(excel_files) == 0:
                            # print(f'Nenhum arquivo encontrado no diretório {month_folder_path}')
                            continue
                        
                        elif len(excel_files) == 1:
                            # print(f'Apenas 1 arquivo encontrado no diretório {month_folder_path}')
                            continue
                        # DataFrame vazio para consolidar os dados
                        consolidated_df = pd.DataFrame()
                        # Iterar sobre os arquivos .xlsx dentro do diretório do mês
                        for file_name in excel_files:                        
                            file_path = os.path.join(month_folder_path, file_name)
                                                       
                            # verificar se a planilha 'RELATÓRIO' ou CONSOLIDADO' existe no arquivo
                            if 'RELATÓRIO' in pd.ExcelFile(file_path).sheet_names:
                                # Ler o arquivo Excel
                                df = pd.read_excel(file_path, sheet_name='RELATÓRIO', engine='openpyxl')
                                # Concatenar o DataFrame
                                consolidated_df = pd.concat([consolidated_df, df], ignore_index=True)
                                             
                            if 'CONSOLIDADO' in pd.ExcelFile(file_path).sheet_names:
                                # Ler o arquivo Excel
                                df = pd.read_excel(file_path, sheet_name='CONSOLIDADO', engine='openpyxl')
                                # Concatenar o DataFrame
                                consolidated_df = pd.concat([consolidated_df, df], ignore_index=True)                                
                        
                        # Salvar o DataFrame consolidado em um novo arquivo na mesma pasta
                        consolidated_file_path = os.path.join(month_folder_path, f'CONSOLIDADO_{month_year}_{subfolder}.xlsx')

                        if os.path.exists(consolidated_file_path):
                            print(f'Arquivo {consolidated_file_path} já existe...')
                        
                        else:
                            try:
                                with pd.ExcelWriter(consolidated_file_path, engine='openpyxl') as writer:
                                    consolidated_df.to_excel(writer, sheet_name="RELATÓRIO", index=False, engine='openpyxl')
                                    
                                    consolidated_df['VALOR TOTAL GERADO'] = consolidated_df['VALOR TOTAL GERADO'].str.replace('.', '').str.replace(',', '.').astype(float)
                                    consolidated_df['VALOR TOTAL FATURAMENTO'] = consolidated_df['VALOR TOTAL FATURAMENTO'].str.replace('.', '').str.replace(',', '.').astype(float)

                                    
                                    # agrupar os valores das colunas "VALOR TOTAL GERADO" e "VALOR TOTAL FATURADO" por "PROJETO", "OBRA" e "CONTRATO LEGADO"
                                    sintese_df = consolidated_df.groupby(['PROJETO', 'OBRA', 'CONTRATO LEGADO'], as_index=False).agg(
                                        {'VALOR TOTAL GERADO': 'sum', 'VALOR TOTAL FATURAMENTO': 'sum'})
                                    
                                    # renomear as colunas
                                    sintese_df = sintese_df.rename(columns={'VALOR TOTAL FATURAMENTO': 'VALOR TOTAL FATURADO'})                    


                                    # formatação da planilha "CONSOLIDADO"
                                    worksheet = writer.sheets['RELATÓRIO']
                                    for column in range(1, worksheet.max_column + 1):
                                        worksheet.column_dimensions[worksheet.cell(row=1, column=column).column_letter].width = 20
                                        worksheet.cell(row=1, column=column).font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
                                        worksheet.cell(row=1, column=column).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                        # alinhar o texto no meio
                                        worksheet.cell(row=1, column=column).alignment = Alignment(horizontal='center', vertical='center', )
                                        worksheet.row_dimensions[1].height = 24

                                    # Configuração para o formato brasileiro
                                    # locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

                                    sintese_df.to_excel(writer, sheet_name='SÍNTESE', index=False)

                                    # Adiciona "TOTAL" abaixo da célula "C"
                                    worksheet = writer.sheets['SÍNTESE']
                                    # print(f'Sheet sintese criada')
                                    worksheet.cell(row=worksheet.max_row + 2, column=4, value='TOTAL')

                                    # negrito na célula "TOTAL"
                                    worksheet.cell(row=worksheet.max_row, column=4).font = Font(bold=True)
                                    worksheet.cell(row=worksheet.max_row, column=5).font = Font(bold=True)

                                                                    
                                    # Soma os valores da coluna "E" (VALOR TOTAL FATURADO) e "D" (VALOR TOTAL GERADO)
                                    # sintese_df['VALOR TOTAL FATURADO'] = sintese_df['VALOR TOTAL FATURADO'].str.replace(',', '.')
                                    # sintese_df['VALOR TOTAL GERADO'] = sintese_df['VALOR TOTAL GERADO'].str.replace(',', '.')
                                    

                                    
                                    total_valor_a_cobrar = sintese_df['VALOR TOTAL GERADO'].sum()
                                    total_valor_total_previo = sintese_df['VALOR TOTAL FATURADO'].sum()
                                    
                                    
                                    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

                                    total_valor_a_cobrar = locale.currency(total_valor_a_cobrar, grouping=True)
                                    total_valor_total_previo = locale.currency(total_valor_total_previo, grouping=True)
                                    
                                    
                                    # formatação da soma dos valores da coluna "D, E"
                                    # total_valor_a_cobrar = "${:,.2f}".format(float(re.sub(r'[^\d.]', '', total_valor_a_cobrar)), grouping=True)
                                    # total_valor_total_previo = "${:,.2f}".format(float(re.sub(r'[^\d.]', '', total_valor_total_previo)), grouping=True)

                                    # negrito na célula "VALOR TOTAL FATURADO"
                                    worksheet.cell(row=worksheet.max_row, column=4).font = Font(bold=True)

                                    # formatação da soma dos valores da coluna "D, E"
                                    # total_valor_a_cobrar = "${:,.2f}".format(float(re.sub(r'[^\d.]', '', total_valor_a_cobrar)), grouping=True)
                                    # total_valor_total_previo = "${:,.2f}".format(float(re.sub(r'[^\d.]', '', total_valor_total_previo)), grouping=True)
                                    # total_valor_a_cobrar = "R${:,.2f}".format(float(total_valor_a_cobrar), grouping=True)
                                    # total_valor_total_previo = "R${:,.2f}".format(float(total_valor_total_previo), grouping=True)

                                    worksheet.cell(row=worksheet.max_row, column=4, value=total_valor_a_cobrar)
                                    worksheet.cell(row=worksheet.max_row, column=5, value=total_valor_total_previo)

                                    # Aplicar cor vermelha ao cabeçalho das colunas A, B, C e D e negrito e tipografia "Alwyn New Light"
                                    for column in 'ABCDE':
                                        header_cell = worksheet[f"{column}1"]
                                        header_cell.font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
                                        header_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                        header_cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                                        # alinhar o texto na esquerda
                                        header_cell.alignment = Alignment(horizontal='left', vertical='center', )

                                    # adicona bordas externas à planilha "SÍNTESE"
                                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                                        for cell in row:
                                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))

                                    # formatar largura das colunas
                                    writer.sheets['SÍNTESE'].column_dimensions['A'].width = 20
                                    writer.sheets['SÍNTESE'].column_dimensions['B'].width = 15
                                    writer.sheets['SÍNTESE'].column_dimensions['C'].width = 31
                                    writer.sheets['SÍNTESE'].column_dimensions['D'].width = 23
                                    writer.sheets['SÍNTESE'].column_dimensions['E'].width = 23
                            except Exception as e:
                                print(f'Erro ao salvar arquivo no arquivo {consolidated_file_path}: {e}')

        except PermissionError as e:
            print(f'Erro de permissão: {e}')
            
                        
    
                          
   

                        


                

            







